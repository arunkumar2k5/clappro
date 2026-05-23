from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, List
from models.comparison import ComponentComparison, MatchStatus
from io import BytesIO

def generate_excel_export(
    parameter_names: List[str],
    base_component: Dict[str, str],
    comparisons: List[ComponentComparison],
    justifications: Dict[str, str],
    recommendation: str,
    confidence_scores: Dict[str, float]
) -> bytes:
    """
    Generate Excel file with comparison results.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Component Comparison"
    
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    
    ws.column_dimensions['A'].width = 25
    for i in range(len(comparisons) + 2):
        ws.column_dimensions[get_column_letter(i + 2)].width = 20
    
    ws['A1'] = "Component"
    ws['A1'].font = bold_font
    ws['B1'] = f"{base_component.get('part_number', 'Base')} (BASE)"
    ws['B1'].font = bold_font
    ws['B1'].fill = blue_fill
    
    col_idx = 3
    for comp in comparisons:
        cell = ws.cell(row=1, column=col_idx)
        cell.value = comp.part_number
        cell.font = bold_font
        col_idx += 1
    
    ws.cell(row=1, column=col_idx).value = "Justification"
    ws.cell(row=1, column=col_idx).font = bold_font
    
    ws['A2'] = "Confidence"
    ws['A2'].font = bold_font
    base_conf = confidence_scores.get('base', 0)
    ws['B2'] = f"{base_conf:.1f}%" if base_conf > 1 else f"{base_conf * 100:.1f}%"
    ws['B2'].alignment = center_align
    
    col_idx = 3
    for comp in comparisons:
        cell = ws.cell(row=2, column=col_idx)
        comp_conf = confidence_scores.get(comp.part_number, 0)
        cell.value = f"{comp_conf:.1f}%" if comp_conf > 1 else f"{comp_conf * 100:.1f}%"
        cell.alignment = center_align
        col_idx += 1
    
    row_idx = 3
    for param_name in parameter_names:
        ws.cell(row=row_idx, column=1).value = param_name
        ws.cell(row=row_idx, column=1).font = bold_font
        
        base_value = base_component.get(param_name, "N/A")
        base_cell = ws.cell(row=row_idx, column=2)
        base_cell.value = base_value
        base_cell.fill = blue_fill
        
        col_idx = 3
        for comp in comparisons:
            param_cell = comp.parameters.get(param_name)
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if param_cell:
                cell.value = param_cell.value
                
                if param_cell.status == MatchStatus.MATCH:
                    cell.fill = green_fill
                    cell.value = f"✓ {param_cell.value}"
                elif param_cell.status == MatchStatus.NO_MATCH:
                    cell.fill = red_fill
                    cell.value = f"✗ {param_cell.value}"
            else:
                cell.value = "N/A"
            
            col_idx += 1
        
        justification_cell = ws.cell(row=row_idx, column=col_idx)
        justification_cell.value = justifications.get(param_name, "")
        
        row_idx += 1
    
    row_idx += 2
    ws.cell(row=row_idx, column=1).value = "RECOMMENDATION"
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    
    ws.cell(row=row_idx + 1, column=1).value = recommendation
    ws.merge_cells(start_row=row_idx + 1, start_column=1, end_row=row_idx + 1, end_column=col_idx)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()
